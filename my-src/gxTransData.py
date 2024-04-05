import pandas as pd
from openpyxl import load_workbook


# 国信证券交易资金流水中根据摘要对成交数量和发生金额的处理系数定义
class SummaryClassifier:
    SUMMARY_CLASSIFICATION = {
        'Tn证券买入': {'成交数量标志': 1, '发生金额标志': 1},
        'Tn证券卖出': {'成交数量标志': -1, '发生金额标志': 1},
        '三方券商签约划入': {'成交数量标志': 1, '发生金额标志': 1},
        '三方券商签约划出': {'成交数量标志': 1, '发生金额标志': 1},
        '三方券商解约划入': {'成交数量标志': 1, '发生金额标志': 1},
        '三方券商解约划出': {'成交数量标志': 1, '发生金额标志': 1},
        '三方银行签约划入': {'成交数量标志': 1, '发生金额标志': 1},
        '三方银行签约划出': {'成交数量标志': 1, '发生金额标志': 1},
        '专项融券回购(日)': {'成交数量标志': 1, '发生金额标志': 1},
        '专项融券回购续约': {'成交数量标志': 1, '发生金额标志': 1},
        '专项融券购回(日)': {'成交数量标志': -1, '发生金额标志': 1},
        '中签缴款': {'成交数量标志': 1, '发生金额标志': 1},
        '保证金理财赎回': {'成交数量标志': -1, '发生金额标志': 1},
        '利息归本': {'成交数量标志': -1, '发生金额标志': 1},
        '利息税代扣': {'成交数量标志': 1, '发生金额标志': 1},
        '基金申购拨出': {'成交数量标志': 1, '发生金额标志': 1},
        '分级基金上折': {'成交数量标志': 1, '发生金额标志': 1},  # 为军工分级手工创建的记录
        '开放基金拆分减股': {'成交数量标志': -1, '发生金额标志': 1},
        '开放基金拆分增股': {'成交数量标志': 1, '发生金额标志': 1},
        '手续费多退少补取': {'成交数量标志': 1, '发生金额标志': 1},
        '投票确认': {'成交数量标志': 0, '发生金额标志': 1},
        '担保品划入': {'成交数量标志': 1, '发生金额标志': 1},  # 这个应该忽略，在股份转出时一起处理了
        '担保品划出': {'成交数量标志': -1, '发生金额标志': 1},
        '新股入帐': {'成交数量标志': 1, '发生金额标志': 1},
        '新股申购': {'成交数量标志': 0, '发生金额标志': 0},  # 申购相当于只是资金冻结，申购中签才扣钱
        '港股通组合费': {'成交数量标志': 1, '发生金额标志': 1},
        '港股通股票买入': {'成交数量标志': 1, '发生金额标志': 1},
        '港股通股票卖出': {'成交数量标志': -1, '发生金额标志': 1},
        '申购中签': {'成交数量标志': 0, '发生金额标志': 1},
        '申购还款': {'成交数量标志': 0, '发生金额标志': 0},  # 申购相当于只是资金冻结，申购中签才扣钱
        '红利入帐0': {'成交数量标志': 0, '发生金额标志': 1},
        '红利入账': {'成交数量标志': 0, '发生金额标志': 1},
        '红利税补扣': {'成交数量标志': 0, '发生金额标志': 1},
        '红股入账': {'成交数量标志': 1, '发生金额标志': 1},
        '股份转入': {'成交数量标志': 1, '发生金额标志': 1},  # 这个应该忽略，在担保品划出时一起处理了
        '股份转出': {'成交数量标志': -1, '发生金额标志': 1},
        '自有资金还融资': {'成交数量标志': 0, '发生金额标志': 1, '银行标志': 1},
        '融入方初始交易': {'成交数量标志': 0, '发生金额标志': 1, '银行标志': 1},  # 这是非融资账户里的，和下一条之间差了个利息
        '融入购回减资金': {'成交数量标志': 0, '发生金额标志': 1, '银行标志': 1},
        '融券回购': {'成交数量标志': 1, '发生金额标志': 1},
        '融券购回': {'成交数量标志': -1, '发生金额标志': 1},
        '融资借款': {'成交数量标志': 0, '发生金额标志': 1, '银行标志': 1},
        '融资利息': {'成交数量标志': 1, '发生金额标志': 1, '银行标志': 0},  # 融资利息不是银行转入转出部分，是账户内消耗
        '融资平仓': {'成交数量标志': -1, '发生金额标志': 1},
        '融资开仓': {'成交数量标志': 1, '发生金额标志': 1},
        '融资还款': {'成交数量标志': 0, '发生金额标志': 1, '银行标志': 1},
        '要约资金': {'成交数量标志': -1, '发生金额标志': 1},  # 爱建要约卖出
        '证券买入': {'成交数量标志': 1, '发生金额标志': 1},
        '证券卖出': {'成交数量标志': -1, '发生金额标志': 1},
        '证券转银行': {'成交数量标志': 1, '发生金额标志': 1, '银行标志': 1},
        '资金冻结': {'成交数量标志': -1, '发生金额标志': 0},
        '资金自动还融资': {'成交数量标志': 0, '发生金额标志': 1, '银行标志': 1},
        '银行转证券': {'成交数量标志': -1, '发生金额标志': 1, '银行标志': 1},
    }

    @staticmethod
    # return volume_flag , amount_flag,bank_flag
    def get_classification(trans_summary):
        """
        Returns the transaction flag and fund change flag for a given summary.
        :param: trans_summary: The summary to look up.
        :return: A tuple containing the transaction flag and fund change flag.
        """
        classification = SummaryClassifier.SUMMARY_CLASSIFICATION.get(trans_summary, None)
        if classification:
            # 判断是否包含'银行标志'这个key
            if '银行标志' in classification:
                bank_flag = classification['银行标志']
            else:
                bank_flag = 0
            return classification['成交数量标志'], classification['发生金额标志'], bank_flag
        else:
            return None, None, None

    @staticmethod
    def get_account_type(currency, is_margin):
        if is_margin == '是':
            account_type = '国信融资账户'
        elif currency == '人民币':
            account_type = '国信账户'
        else:
            account_type = '国信B股'
        return account_type


# 国信证券交易资金流水中的初始持仓
class AccountSummary:
    ACCOUNT_SUMMARY_FILE = 'analyze_summary.xlsx'
    # 每只股票在初始20070507时点的初始持仓股数
    INIT_HOLDINGS = {  # 持股成本价以20070507时的不复权股价为计算依据
        '600161': {'交收日期': '20070507', '账户': '国信账户', '证券名称': '天坛生物', '持股数量': 12000, '持股成本价': 20.42},
        '600677': {'交收日期': '20070507', '账户': '国信账户', '证券名称': '航天通信', '持股数量': 18000, '持股成本价': 32.7},
        '580006': {'交收日期': '20070507', '账户': '国信账户', '证券名称': '雅戈QCB1', '持股数量': 1, '持股成本价': 17.1},
        '900947': {'交收日期': '20070507', '账户': '国信B股', '证券名称': '振华B股', '持股数量': 9900, '持股成本价': 1.6},
        '900932': {'交收日期': '20070507', '账户': '国信B股', '证券名称': '陆家B股', '持股数量': 100, '持股成本价': 2.5}
    }

    # 账户在初始20070507时点的资金余额
    INIT_CAPITAL = {  # 20070510时的初始资金余额（不含股票）
        '国信B股': {'交收日期': '20070507', '累计净转入资金': 25000.0, '资金余额': 0.0},
        '国信账户': {'交收日期': '20070507', '累计净转入资金': 300000.0, '资金余额': 45046.37},  # 这是从后续交易中倒推算出来的
        '国信融资账户': {'交收日期': '20070507', '累计净转入资金': 0.0, '资金余额': 0.0},
    }

    def __init__(self):
        self.stockhold_history = pd.DataFrame()  # 每日持仓历史
        self.balance_history = pd.DataFrame()  # 每日资金余额历史
        self.stock_profit_history = pd.DataFrame()  # 股票实现盈亏历史（不包含浮盈）
        self.stockhold_record = None
        self.balance_record = None

    # 根据传入的起始日期设定当日的初始持仓数据
    def init_start_holdings(self, start_date=None):
        self.stockhold_record = AccountSummary.init_stockhold_record(start_date)
        self.balance_record = AccountSummary.init_balance_record(start_date)
        return self.stockhold_record, self.balance_record

    @staticmethod
    def init_stockhold_record(start_date=None):
        # 初始化每日持股数据DataFrame
        stockhold_data = []
        if start_date is None:  # 从2007年开始的全量数据分析
            for code, details in AccountSummary.INIT_HOLDINGS.items():
                stockhold_data.append({
                    '交收日期': pd.to_datetime(details['交收日期'], format='%Y%m%d'),
                    '账户类型': details['账户'],
                    '证券代码': code,
                    '证券名称': details['证券名称'],
                    '持股数量': details['持股数量'],
                    '持股成本': details['持股数量'] * details['持股成本价'],
                    '当日市值': 0.0,
                    '浮动盈亏': 0.0
                })
        else:  # 从特定日期开始的增量数据分析
            stockhold_history = AccountSummary.load_stockhold_from_file()
            # 找到 start_date 前一天的持仓记录
            prev_day = start_date - pd.Timedelta(days=1)
            stockhold_data = stockhold_history[stockhold_history['交收日期'] == prev_day]
        return pd.DataFrame(stockhold_data)

    @staticmethod
    def init_balance_record(start_date=None):
        # 初始化每日资金余额数据DataFrame
        balance_data = []
        if start_date is None:  # 从2007年开始的全量数据分析
            for account, details in AccountSummary.INIT_CAPITAL.items():
                balance_data.append({
                    '交收日期': pd.to_datetime(details['交收日期'], format='%Y%m%d'),
                    '账户类型': account,
                    '累计净转入资金': details['累计净转入资金'],
                    '资金余额': details['资金余额'],
                    '记录账户余额': 0.0,  # 交易文件中记录的账户余额
                    '校验差异': 0.0,  # 校验数据
                    '盈亏': 0.0, #盈亏放前面，以免analyzeAccountProfit时错位
                    '当日市值': 0.0,
                })
        else:  # 从特定日期开始的增量数据分析
            balance_history = AccountSummary.load_balance_from_file()
            # 找到 start_date 前一天的持仓记录
            prev_day = start_date - pd.Timedelta(days=1)
            balance_data = balance_history[balance_history['交收日期'] == prev_day]
        return pd.DataFrame(balance_data)

    # 从 'analyze_summary.xlsx' 文件中加载历史的持仓记录
    @staticmethod
    def load_stockhold_from_file():
        return pd.read_excel(AccountSummary.ACCOUNT_SUMMARY_FILE, sheet_name="股票持仓历史", header=0,
                             dtype={'证券代码': str})

    # 从 'analyze_summary.xlsx' 文件中加载历史的账户余额记录
    @staticmethod
    def load_balance_from_file():
        return pd.read_excel(AccountSummary.ACCOUNT_SUMMARY_FILE, sheet_name="账户余额历史", header=0)

    # 加载每日持股记录和账户数据
    def load_account_summaries(self, start_date=None):
        """
        Load data from 'analyze_summary.xlsx'.
        If start_date is provided, only load data from that date onwards.
        """
        stock_holding_records = self.load_stockhold_from_file()
        account_balance_records = self.load_balance_from_file()

        if start_date:
            stock_holding_records = stock_holding_records[stock_holding_records['交收日期'] >= start_date]
            account_balance_records = account_balance_records[account_balance_records['交收日期'] >= start_date]

        return stock_holding_records, account_balance_records

    def add_to_history(self, new_balance_row, new_holdings):
        self.balance_history = pd.concat([self.balance_history, new_balance_row], ignore_index=True)
        self.stockhold_history = pd.concat([self.stockhold_history, new_holdings], ignore_index=True)

    def add_to_stock_profit_history(self, new_profit_records):
        # 去除掉持股成本为0的，并把持股成本变符号
        new_profit_records = new_profit_records.drop(new_profit_records[new_profit_records['持股成本'] == 0].index)
        new_profit_records['持股成本'] *= -1
        self.stock_profit_history = pd.concat([self.stock_profit_history, new_profit_records], ignore_index=True)

    def save_account_history(self, start_date=None):
        # 个股盈亏数据的重命名
        self.stock_profit_history.rename(columns={'持股成本': '实现盈亏'}, inplace=True)
        self.stock_profit_history = self.stock_profit_history[['交收日期', '账户类型', '证券代码', '证券名称', '持股数量', '实现盈亏']]

        # 创建一个Excel文件
        if start_date:  # 增量模式
            with pd.ExcelWriter(AccountSummary.ACCOUNT_SUMMARY_FILE, engine='openpyxl',
                                mode='a', if_sheet_exists='overlay') as writer:
                # 将数据写入Excel文件的不同sheet
                self.append_data_to_sheet(writer, '账户余额历史', self.balance_history, start_date)
                self.append_data_to_sheet(writer, '股票持仓历史', self.stockhold_history, start_date)
                self.append_data_to_sheet(writer, '个股盈亏历史', self.stock_profit_history, start_date)

        else:  # 全量从2007年开始分析的模式
            # 创建一个新的Excel文件
            with pd.ExcelWriter(AccountSummary.ACCOUNT_SUMMARY_FILE, mode='w', engine='openpyxl') as writer:
                # 将数据写入Excel文件的不同sheet
                # Format the numbers uniformly as '###,###,###.##'
                self.balance_history.to_excel(writer, sheet_name='账户余额历史', index=False)
                self.stockhold_history.to_excel(writer, sheet_name='股票持仓历史', index=False)
                self.stock_profit_history.to_excel(writer, sheet_name='个股盈亏历史', index=False)

        self.format_account_summary_file()

    # 设置格式
    @staticmethod
    def format_account_summary_file():
        # 打开已创建的Excel文件设置格式
        wb = load_workbook(AccountSummary.ACCOUNT_SUMMARY_FILE)
        # 逐一选择三个不同的sheet并设置格式
        for sheet_name in ['账户余额历史', '股票持仓历史', '个股盈亏历史']:
            ws = wb[sheet_name]
            # 设置日期格式为YYYY/MM/DD
            for cell in ws['A']:
                cell.number_format = 'YYYY/MM/DD'
        # 保存更改后的Excel文件
        wb.save(AccountSummary.ACCOUNT_SUMMARY_FILE)

    # 在已有的excel文件指定sheet日期之后追加新的增量数据
    @staticmethod
    def append_data_to_sheet(writer, sheet_name, df_new_data, start_date):
        # 确保新数据只有start_date之后的
        df_new_data = df_new_data[df_new_data['交收日期'] >= start_date]
        # 加载已有的数据
        df_old_data = pd.DataFrame(pd.read_excel(AccountSummary.ACCOUNT_SUMMARY_FILE, sheet_name=sheet_name))
        df_old_data['交收日期'] = pd.to_datetime(df_old_data['交收日期'])
        # 将数据df_new_data写入excel中的sheet表,从start_date之后的第一个行开始覆盖/追加写：
        skip_rows = df_old_data[df_old_data['交收日期'] < start_date].shape[0]
        df_new_data.to_excel(writer, sheet_name=sheet_name, startrow=skip_rows + 1, index=False, header=False)


# Example usage
if __name__ == "__main__":
    # 使用示例
    account_summary = AccountSummary()
    # 初始化持股数据, 初始化资金余额数据:
    startDate = pd.to_datetime('20231124', format='%Y%m%d')
    init_stockhold, init_balance = account_summary.init_start_holdings(startDate)
    print(f"{startDate}持股数据:")
    print(init_stockhold)
    print("\n{start_date}资金余额数据:")
    print(init_balance)
